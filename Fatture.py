# -*- coding: utf-8 -*-
"""
Created on Mon Sep 21 12:09:59 2020

@author: user
"""
# Only works on Windows

import openpyxl as xl
from datetime import date
import win32com.client
from pywintypes import com_error

class InvoiceGenerator:

    def __init__(self, descriptions_sheet, import_sheet, export_sheet, export_path):
        self.descriptions = xl.load_workbook(descriptions_sheet, read_only=True,
                                             data_only=True).active
        self.data = xl.load_workbook(import_sheet, read_only=True,
                                     data_only=True).active
        self.form = xl.load_workbook(export_sheet, read_only=False,
                                     data_only=True).active
        self.excel = win32com.client.Dispatch("Excel.Application")
        self.excel.Visible = False
        self.xp_sheet = export_sheet
        self.pdf_path = export_path
        self.orders = []
        
    def get_orders(self):
        today = date.today()
        invoice_num = 1
        for row in self.data.iter_rows(min_row=2, max_row=self.data.max_row,
                    min_col=1, max_col=self.data.max_column, values_only=True):
            # Create order - store values in key-value pairs
                order = {"Buyer Name":row[11],
                         "Shipping Address":row[25],
                         "Shipping Details": f"{row[28]}, {row[29]} {row[30]}",
                         "Invoice num": invoice_num, # Set initial invoice number
                         "Emission Date": today.strftime("%d/%m/%Y"),
                         "Order ID":row[0],
                         "Items Qty": int(1),
                         "Product Code":row[13],
                         "Unit Price":float(row[17]),
                         "Total Price":float(row[17]),
                         "Date String": today.strftime("%B %d, %Y")}
            
                # Get all relevant values from rows with matching order ID
                for row in self.descriptions.iter_rows(min_row=2,
                            max_row=self.descriptions.max_row, min_col=1,
                            max_col=self.descriptions.max_column, values_only=True):
                    if order["Product Code"] == row[0]:
                        order.update({"Description": row[3],
                                      "HTS Heading": row[1],
                                      "FDA Code": row[2]})
                        break
                self.orders.append(order)
                invoice_num += 1
                
    def merge_same_id(self):
        for i in reversed(range(len(self.orders))):
            if self.orders[i]["Order ID"] == self.orders[i-1]["Order ID"]:
                self.orders[i]["Total Price"] += self.orders[i-1]["Unit Price"]
                self.orders[i]["Items Qty"] += 1
                self.orders.remove(self.orders[i-1])
                
    def fill_form_and_export(self):
        for order in self.orders:
            # Send order to correct cells in excel output file with default template (do not overwrite)
            cells = ["A14", "A15", "A16", "C22", "C23", "C24", "A30", "B30", "I30", "J30", "B47", "C30", "E30", "G30"]
            i = 0
            for key in order:
                self.form[cells[i]] = order[key]
                i += 1
            # Save excel file as pdf named with order ID
            try:
                formXl = self.excel.Workbooks.Open(self.export_sheet)
                print("Start conversion to PDF: {}".format(order["Buyer Name"]))
                formXl.ActiveSheet.ExportAsFixedFormat(0, f'{self.pdf_path} {order["Order ID"]}')
            except com_error as e:
                print("Failed: " + e)
            finally:
                formXl.Close()
            # Checkpoint
            self.form.save(r"C:\Users\t_man\Documents\ThatsArte Fatture XL\Output Form.xlsx")
            # Overwrites the excel file by looping through again
            
    def quit_excel(self):
        self.excel.Quit()
        
    def print_orders(self):
        for order in self.orders:
            print(f"Order Number --- {order[Invoice num]}")
            for key, value in order.items():
                print(f"\t{key}: {value}")
            print("\n")
